"""Tests covering the coercion and extraction helpers in :mod:`app`."""

from __future__ import annotations

import logging
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from excel_io_pkg import excel_io
from extract.models import ExtractionResult
from extract.numeric import normalize_numeric_string

import app


class DummyEngine:
    def process(self, path: Path):  # pragma: no cover - trivial stub
        return SimpleNamespace(path=path)


class DummyIndex:
    def __init__(self, entry):
        self._entry = entry

    def get_latest(self, mb: str, form_type: str, preferred_period: str):
        return self._entry


def test_coerce_extraction_result_collects_messages():
    result = ExtractionResult(field_name="revenue", value=101)
    result.add_warning("low_conf", "Low OCR confidence", confidence=0.42)
    result.add_error("missing_anchor", "Anchor not found", anchor="пословни приходи")

    value, notes, meta = app._coerce_extractor_result(result)

    assert value == 101
    assert meta == {"has_errors": True, "has_warnings": True, "has_value": True}
    assert any(note.startswith("ERROR missing_anchor") for note in notes)
    assert any("WARNING low_conf" in note for note in notes)
    assert any("anchor=пословни приходи" in note for note in notes)


def test_main_reports_xls_conversion_requirement(monkeypatch: pytest.MonkeyPatch, tmp_path, caplog):
    xls_path = tmp_path / "input.xls"
    xls_path.write_text("dummy")

    monkeypatch.setattr(
        app,
        "parse_args",
        lambda argv=None: SimpleNamespace(excel=str(xls_path), year=None, force=False, debug=False),
    )
    monkeypatch.setattr(app, "load_config", lambda path: {"log_path": tmp_path / "log.log"})

    def fake_configure_logging(*, debug: bool, log_path: Path):
        logger = logging.getLogger("finstat-test")
        logger.setLevel(logging.DEBUG if debug else logging.INFO)
        logger.propagate = True
        return logger

    monkeypatch.setattr(app, "configure_logging", fake_configure_logging)
    monkeypatch.setattr(app, "configure_dependencies", lambda config, logger: None)
    def raise_invalid(*args, **kwargs):
        raise InvalidFileException("legacy")

    monkeypatch.setattr(app.excel_io, "load_workbook", raise_invalid)

    with caplog.at_level(logging.ERROR):
        exit_code = app.main([])

    assert exit_code == 2
    assert any("convert it to .xlsx" in record.getMessage() for record in caplog.records)


def test_extract_single_field_marks_missing_on_error(monkeypatch: pytest.MonkeyPatch):
    entry = SimpleNamespace(path=Path("dummy.pdf"))
    index = DummyIndex(entry)
    engine = DummyEngine()

    extraction_result = ExtractionResult(field_name="revenue", value=500)
    extraction_result.add_error("no_value", "Value not detected")

    def fake_resolve(form_type: str, field: str):
        assert form_type == "bu"
        assert field == "revenue"
        return lambda ocr_result, **kwargs: extraction_result

    monkeypatch.setattr(app, "_resolve_extractor", fake_resolve)

    value, notes, missing = app._extract_single_field(
        mb="12345678",
        form_type="bu",
        field="revenue",
        index=index,
        engine=engine,
        year_preference="current",
        config={},
        logger=logging.getLogger(__name__),
    )

    assert value is None
    assert missing == ["revenue"]
    assert any(note.startswith("ERROR no_value") for note in notes)


def test_extract_multiple_fields_skips_failed_values(monkeypatch: pytest.MonkeyPatch):
    entry = SimpleNamespace(path=Path("dummy.pdf"))
    index = DummyIndex(entry)
    engine = DummyEngine()

    asset_result = ExtractionResult(field_name="assets", value=100)
    asset_result.add_error("not_found", "Assets row not found")

    loss_result = ExtractionResult(field_name="capital_loss", value=200)
    loss_result.add_warning("low_conf", "Confidence below threshold", confidence=0.3)

    def fake_resolve(form_type: str, field: str):
        mapping = {
            "assets": asset_result,
            "capital_loss": loss_result,
        }

        def _extractor(ocr_result, **kwargs):
            return mapping[field]

        return _extractor

    monkeypatch.setattr(app, "_resolve_extractor", fake_resolve)

    values, notes, missing = app._extract_multiple_fields(
        mb="12345678",
        form_type="bs",
        fields=["assets", "capital_loss"],
        index=index,
        engine=engine,
        year_preference="current",
        config={},
        logger=logging.getLogger(__name__),
    )

    assert values == {"capital_loss": 200}
    assert missing == ["assets"]
    assert any(note.startswith("ERROR not_found") for note in notes)
    assert any("WARNING low_conf" in note for note in notes)


def test_empty_numeric_string_persists_zero_in_excel():
    parse_result = normalize_numeric_string("   ")
    extraction_result = ExtractionResult(field_name="revenue")
    extraction_result.value = parse_result.value
    extraction_result.raw_text = "   "
    extraction_result.normalized_text = parse_result.normalized_text

    value, notes, meta = app._coerce_extractor_result(extraction_result)

    assert value == 0
    assert notes == []
    assert meta == {"has_errors": False, "has_warnings": False, "has_value": True}
    assert extraction_result.success

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.cell(row=1, column=1, value="Матични број")
    column_map = excel_io.ensure_result_columns(sheet)
    row_index = 2
    sheet.cell(row=row_index, column=1, value="12345678")

    updates = excel_io.write_result_row(sheet, row_index, {"revenue": value})
    assert updates["revenue"]
    assert sheet.cell(row=row_index, column=column_map["revenue"]).value == 0

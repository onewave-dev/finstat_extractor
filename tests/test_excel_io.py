"""Tests for :mod:`io.excel_io`."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
spec = importlib.util.spec_from_file_location(
    "project_excel_io", ROOT / "io" / "excel_io.py"
)
excel_io = importlib.util.module_from_spec(spec)
assert spec.loader is not None
sys.modules[spec.name] = excel_io
spec.loader.exec_module(excel_io)


def _create_sheet():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.cell(row=1, column=1, value="Матични број")
    return workbook, sheet


def test_ensure_result_columns_populates_headers():
    workbook, sheet = _create_sheet()

    mapping = excel_io.ensure_result_columns(sheet)

    assert sheet["G1"].value == "Пословни приходи (000 РСД)"
    assert sheet["H1"].value == "Укупна актива (000 РСД)"
    assert sheet["I1"].value == "Губитак изнад висине капитала (000 РСД)"
    assert mapping["revenue"] < mapping["assets"] < mapping["capital_loss"]


def test_iter_maticni_broj_normalizes_values():
    workbook, sheet = _create_sheet()
    sheet.cell(row=2, column=1, value="  12-34  56  ")
    sheet.cell(row=3, column=1, value=" ")

    column = excel_io.find_maticni_broj_column(sheet)
    rows = list(excel_io.iter_maticni_broj(sheet, column))

    assert rows[0].mb == "123456"
    assert rows[0].row_index == 2
    assert rows[1].mb == ""


def test_write_result_row_respects_overwrite_flag():
    workbook, sheet = _create_sheet()
    excel_io.ensure_result_columns(sheet)
    sheet.cell(row=2, column=7, value="existing")

    updates = excel_io.write_result_row(sheet, 2, {"revenue": 200})
    assert not updates["revenue"]
    assert sheet.cell(row=2, column=7).value == "existing"

    updates = excel_io.write_result_row(sheet, 2, {"revenue": 200}, overwrite_nonempty=True)
    assert updates["revenue"]
    assert sheet.cell(row=2, column=7).value == 200
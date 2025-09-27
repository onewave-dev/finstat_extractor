"""Excel I/O utilities used by :mod:`app`.

The helpers provided here are intentionally small and predictable so that the
CLI can focus on orchestration:

* locate the column that contains ``"Матични број"`` values and normalise
  them to digit-only strings;
* ensure that result columns **G/H/I** exist with the canonical headers from
  the project specification;
* write results back to the worksheet while respecting the
  ``overwrite_nonempty`` flag.
"""

from __future__ import annotations

import re
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Mapping, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Column metadata
# ---------------------------------------------------------------------------


#: Mapping describing the output columns that must exist in the Excel sheet.
#: The order is significant and matches the specification from ``AGENTS.md``.
RESULT_COLUMNS: "OrderedDict[str, tuple[str, str]]" = OrderedDict(
    [
        ("revenue", ("G", "Пословни приходи (000 РСД)")),
        ("assets", ("H", "Укупна актива (000 РСД)")),
        ("capital_loss", ("I", "Губитак изнад висине капитала (000 РСД)")),
    ]
)


#: Candidate headers for the "Матични број" column.  They are normalised by
#: :func:`_normalise_header` before comparison so casing, spacing and
#: punctuation differences are tolerated.
MB_HEADER_CANDIDATES: Iterable[str] = (
    "матични број",
    "матични бр",
    "мат. број",
    "mat. broj",
    "maticni broj",
    "matični broj",
    "maticni br",
    "matični br",
    "mb",
)


@dataclass
class ExcelRow:
    """Represents a single data row inside the Excel sheet."""

    row_index: int
    raw_value: Optional[str]
    mb: str


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------


def load_workbook_with_sheet(path: Path) -> tuple[Workbook, Worksheet]:
    """Load ``path`` and return the workbook together with the active sheet."""

    workbook = load_workbook(filename=str(path))
    worksheet = workbook.active
    return workbook, worksheet


def find_maticni_broj_column(sheet: Worksheet, *, header_row: int = 1) -> int:
    """Locate the column containing the "Матични број" header.

    Raises ``ValueError`` if the column cannot be located.
    """

    target_headers = {_normalise_header(value) for value in MB_HEADER_CANDIDATES}

    for cell in sheet[header_row]:
        header = _normalise_header(cell.value)
        if header and header in target_headers:
            return cell.column

    raise ValueError("Column 'Матични број' not found in the header row")


def iter_maticni_broj(
    sheet: Worksheet, column: int, *, start_row: int = 2
) -> Iterator[ExcelRow]:
    """Yield :class:`ExcelRow` objects for each data row in ``sheet``."""

    for row_index in range(start_row, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=column)
        raw_value = cell.value
        mb = normalise_maticni_broj(raw_value)
        yield ExcelRow(row_index=row_index, raw_value=_coerce_to_str(raw_value), mb=mb)


def ensure_result_columns(sheet: Worksheet, *, header_row: int = 1) -> Dict[str, int]:
    """Ensure output columns G/H/I exist with the correct headers.

    Returns a mapping from logical field name
    (``"revenue"``/``"assets"``/``"capital_loss"``) to the 1-based column
    index for convenience.
    """

    column_indices: Dict[str, int] = {}
    for key, (column_letter, header) in RESULT_COLUMNS.items():
        index = column_index_from_string(column_letter)
        cell = sheet.cell(row=header_row, column=index)
        if not _cell_has_value(cell.value) or _normalise_header(cell.value) != _normalise_header(header):
            cell.value = header
        column_indices[key] = index
    return column_indices


def read_rows(path: Path, *, header_row: int = 1) -> tuple[Workbook, Worksheet, List[ExcelRow]]:
    """Load ``path`` and return workbook, sheet and the Matični broj rows."""

    workbook, worksheet = load_workbook_with_sheet(path)
    column = find_maticni_broj_column(worksheet, header_row=header_row)
    rows = list(iter_maticni_broj(worksheet, column, start_row=header_row + 1))
    return workbook, worksheet, rows


def write_result_row(
    sheet: Worksheet,
    row_index: int,
    values: Mapping[str, Optional[object]],
    *,
    overwrite_nonempty: bool = False,
) -> Dict[str, bool]:
    """Write ``values`` to ``row_index`` while honouring ``overwrite_nonempty``.

    Returns a mapping where each key corresponds to the logical field name and
    the boolean value indicates whether the cell was updated.
    """

    updates: Dict[str, bool] = {}
    for key, (column_letter, _) in RESULT_COLUMNS.items():
        column_index = column_index_from_string(column_letter)
        cell = sheet.cell(row=row_index, column=column_index)
        value = values.get(key)

        if value is None and not overwrite_nonempty and not _cell_has_value(cell.value):
            updates[key] = False
            continue

        if not overwrite_nonempty and _cell_has_value(cell.value):
            updates[key] = False
            continue

        cell.value = value
        updates[key] = True

    return updates


def get_existing_values(sheet: Worksheet, row_index: int) -> Dict[str, Optional[object]]:
    """Return the current values from the result columns for ``row_index``."""

    existing: Dict[str, Optional[object]] = {}
    for key, (column_letter, _) in RESULT_COLUMNS.items():
        column_index = column_index_from_string(column_letter)
        existing[key] = sheet.cell(row=row_index, column=column_index).value
    return existing


def save_workbook(workbook: Workbook, path: Path) -> None:
    """Persist the workbook back to disk."""

    workbook.save(str(path))


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------


def normalise_maticni_broj(value: Optional[object]) -> str:
    """Normalise the Matični број value to a string containing only digits."""

    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits.strip()


def _normalise_header(value: Optional[object]) -> str:
    if value is None:
        return ""
    text = str(value).lower()
    text = re.sub(r"[\s\.:;\-]+", "", text)
    return text


def _cell_has_value(value: Optional[object]) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _coerce_to_str(value: Optional[object]) -> Optional[str]:
    if value is None:
        return None
    return str(value)


__all__ = [
    "ExcelRow",
    "RESULT_COLUMNS",
    "ensure_result_columns",
    "find_maticni_broj_column",
    "get_existing_values",
    "iter_maticni_broj",
    "load_workbook_with_sheet",
    "normalise_maticni_broj",
    "read_rows",
    "save_workbook",
    "write_result_row",
]
